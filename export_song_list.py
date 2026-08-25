from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from song_database import USDX_CONFIG, SongDatabase


DEFAULT_OUTPUT = Path("song_list.xlsx")

SHEET_NAME = "Songs"
HEADERS = ("Artist", "Title")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Export the USDX song database to an xlsx file with Artist and "
            "Title columns, sorted alphabetically by Artist. Re-run any "
            "time songs are added or removed to refresh the file - it is "
            "fully regenerated from the current song directories each run."
        ),
    )

    parser.add_argument(
        "-o", "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Path to the xlsx file to write (default: {DEFAULT_OUTPUT}).",
    )
    parser.add_argument(
        "-c", "--config",
        type=Path,
        default=USDX_CONFIG,
        help="Path to the USDX config.ini to read song directories from "
             "(ignored if --song-dir is given).",
    )
    parser.add_argument(
        "-d", "--song-dir",
        type=Path,
        action="append",
        dest="song_dirs",
        metavar="DIR",
        help="Song directory to scan. Can be given multiple times. "
             "If provided, song directories are not read from config.ini.",
    )

    return parser.parse_args()


def build_rows(database: SongDatabase) -> list[tuple[str, str]]:
    unique = {(song.artist, song.title) for song in database.songs}

    return sorted(
        unique,
        key=lambda row: (row[0].casefold(), row[1].casefold()),
    )


def print_duplicates(database: SongDatabase) -> None:
    counts: dict[tuple[str, str], int] = {}
    for song in database.songs:
        key = (song.artist, song.title)
        counts[key] = counts.get(key, 0) + 1

    duplicates = sorted(
        (key for key, count in counts.items() if count > 1),
        key=lambda row: (row[0].casefold(), row[1].casefold()),
    )

    if not duplicates:
        print("No duplicate songs found")
        return

    print(f"Found {len(duplicates)} duplicate songs:")
    for artist, title in duplicates:
        print(f"  {artist} - {title} ({counts[(artist, title)]}x)")


def write_xlsx(rows: list[tuple[str, str]], output: Path) -> None:
    if output.exists():
        workbook = load_workbook(output)
        if SHEET_NAME in workbook.sheetnames:
            del workbook[SHEET_NAME]
        sheet = workbook.create_sheet(SHEET_NAME, 0)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME

    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(row)

    sheet.freeze_panes = "A2"
    autofit_columns(sheet)

    workbook.save(output)


def autofit_columns(sheet) -> None:
    # Character counts overstate width vs. Excel's actual autofit, since
    # Excel sizes to rendered pixel width in a proportional font while a
    # flat "+2" padding assumption overshoots most on short columns.
    for column_cells in sheet.columns:
        longest = max(len(str(cell.value)) for cell in column_cells)
        column_letter = column_cells[0].column_letter
        sheet.column_dimensions[column_letter].width = longest * 0.9 + 1


def main() -> None:
    args = parse_args()

    database = SongDatabase(config_path=args.config, song_dirs=args.song_dirs)
    database.load()

    print_duplicates(database)

    rows = build_rows(database)
    write_xlsx(rows, args.output)

    print(f"Wrote {len(rows)} songs to {args.output}")


if __name__ == "__main__":
    main()
